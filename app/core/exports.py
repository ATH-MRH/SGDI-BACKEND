"""Utilitaires partagés pour l'export Excel et PDF."""
import io
from datetime import date
from typing import Any

from fastapi import HTTPException
from fastapi.responses import StreamingResponse


# ── Excel (openpyxl) ──────────────────────────────────────────────────────────

def excel_response(filename: str, sheets: list[dict]) -> StreamingResponse:
    """
    sheets = [{"title": str, "headers": [...], "rows": [[...], ...], "totals": [...] | None}]
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError as exc:
        raise HTTPException(status_code=501, detail="Export Excel indisponible sur ce déploiement léger") from exc

    wb = Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill("solid", fgColor="043970")
    TOTAL_FILL = PatternFill("solid", fgColor="E8F0FB")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    TOTAL_FONT = Font(bold=True, size=10)
    THIN = Side(style="thin", color="D1D5DB")
    BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))

    for sheet_def in sheets:
        ws = wb.create_sheet(title=sheet_def["title"][:31])
        headers = sheet_def["headers"]
        rows = sheet_def["rows"]
        totals = sheet_def.get("totals")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")
                if isinstance(value, (int, float)) and col_idx > 1:
                    cell.number_format = '#,##0.00'

        if totals:
            total_row = len(rows) + 2
            for col_idx, value in enumerate(totals, 1):
                cell = ws.cell(row=total_row, column=col_idx, value=value)
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF (reportlab) ───────────────────────────────────────────────────────────

BLUE = (4/255, 57/255, 112/255)
LIGHT = (0.95, 0.97, 1.0)
BLACK = (0, 0, 0)
GRAY = (0.4, 0.4, 0.4)
RED = (0.8, 0.1, 0.1)
GREEN = (0.0, 0.5, 0.2)


def pdf_document(filename: str, title: str, meta: list[tuple[str, str]], columns: list[tuple[str, float]], rows: list[list], totals: dict[str, Any] | None = None, notes: str | None = None) -> StreamingResponse:
    """
    columns = [(header, width_pct), ...]  width_pct in 0..1, sum should be ~1
    rows = [[cell, ...], ...]
    totals = {"HT": 1000, "TVA": 190, "TTC": 1190}
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.utils import simpleSplit
    except ImportError as exc:
        raise HTTPException(status_code=501, detail="Export PDF indisponible sur ce déploiement léger") from exc

    buf = io.BytesIO()
    PAGE_W, PAGE_H = A4
    MARGIN = 20 * mm
    CONTENT_W = PAGE_W - 2 * MARGIN
    ROW_H = 8 * mm
    HEADER_H = 10 * mm

    c = rl_canvas.Canvas(buf, pagesize=A4)
    y = PAGE_H - MARGIN

    # ── En-tête ──
    c.setFillColorRGB(*BLUE)
    c.rect(MARGIN, y - 18*mm, CONTENT_W, 18*mm, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(MARGIN + 6*mm, y - 12*mm, title)
    c.setFont("Helvetica", 9)
    c.drawRightString(PAGE_W - MARGIN, y - 12*mm, date.today().strftime("%d/%m/%Y"))
    y -= 22*mm

    # ── Méta ──
    c.setFillColorRGB(*BLACK)
    c.setFont("Helvetica-Bold", 9)
    for label, value in meta:
        c.setFillColorRGB(*GRAY)
        c.drawString(MARGIN, y, label)
        c.setFillColorRGB(*BLACK)
        c.drawString(MARGIN + 40*mm, y, str(value or ""))
        y -= 6*mm
    y -= 4*mm

    # ── En-tête colonnes ──
    col_widths = [w * CONTENT_W for _, w in columns]
    c.setFillColorRGB(*BLUE)
    c.rect(MARGIN, y - HEADER_H, CONTENT_W, HEADER_H, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 8)
    x = MARGIN
    for (header, _), col_w in zip(columns, col_widths):
        c.drawString(x + 2*mm, y - 6.5*mm, header[:25])
        x += col_w
    y -= HEADER_H

    # ── Lignes ──
    c.setFont("Helvetica", 8)
    for i, row in enumerate(rows):
        if y < MARGIN + ROW_H * 3:
            c.showPage()
            y = PAGE_H - MARGIN
        if i % 2 == 0:
            c.setFillColorRGB(*LIGHT)
            c.rect(MARGIN, y - ROW_H, CONTENT_W, ROW_H, fill=1, stroke=0)
        c.setFillColorRGB(*BLACK)
        x = MARGIN
        for cell, col_w in zip(row, col_widths):
            text = str(cell if cell is not None else "")
            c.drawString(x + 2*mm, y - 5.5*mm, text[:35])
            x += col_w
        c.setStrokeColorRGB(0.9, 0.9, 0.9)
        c.line(MARGIN, y - ROW_H, MARGIN + CONTENT_W, y - ROW_H)
        y -= ROW_H

    # ── Totaux ──
    if totals:
        y -= 4*mm
        c.setFillColorRGB(*LIGHT)
        total_block_h = len(totals) * 7*mm + 4*mm
        c.rect(MARGIN + CONTENT_W * 0.55, y - total_block_h, CONTENT_W * 0.45, total_block_h, fill=1, stroke=0)
        tx = MARGIN + CONTENT_W * 0.58
        ty = y - 6*mm
        for label, value in totals.items():
            c.setFillColorRGB(*GRAY)
            c.setFont("Helvetica", 8)
            c.drawString(tx, ty, label)
            c.setFillColorRGB(*BLACK)
            c.setFont("Helvetica-Bold", 9)
            c.drawRightString(MARGIN + CONTENT_W - 2*mm, ty, f"{value:,.2f} DZD" if isinstance(value, (int, float)) else str(value))
            ty -= 7*mm

    # ── Notes ──
    if notes:
        y -= (total_block_h + 6*mm) if totals else 4*mm
        c.setFillColorRGB(*GRAY)
        c.setFont("Helvetica-Oblique", 8)
        for line in simpleSplit(notes, "Helvetica-Oblique", 8, CONTENT_W):
            c.drawString(MARGIN, y, line)
            y -= 5*mm

    # ── Pied de page ──
    c.setFillColorRGB(*GRAY)
    c.setFont("Helvetica", 7)
    c.drawCentredString(PAGE_W / 2, MARGIN / 2, f"Document généré automatiquement — {date.today().strftime('%d/%m/%Y')}")

    c.save()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
